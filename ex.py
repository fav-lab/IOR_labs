from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

excel_path = "grouped_statewise_table_formatted.xlsx"
wb = load_workbook(excel_path)
ws = wb.active

# ====== Fills ======
header_fill = PatternFill("solid", fgColor="F39C12")       # Orange
state_fill = PatternFill("solid", fgColor="145A32")        # Dark Green
light_green_fill = PatternFill("solid", fgColor="D5F5E3")  # Light Green

# ====== Fonts ======
header_font = Font(bold=True, color="000000")         # Black Bold
state_font = Font(bold=True, color="FFFFFF")          # White Bold
white_bold_font = Font(bold=True, color="FFFFFF")     # White Bold for numbers/text

# ====== Alignment ======
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ====== Black Medium Border ======
black_border = Border(
    left=Side(style='medium', color='000000'),
    right=Side(style='medium', color='000000'),
    top=Side(style='medium', color='000000'),
    bottom=Side(style='medium', color='000000')
)

# ====== Set Column Widths ======
for col in range(1, 6):
    ws.column_dimensions[get_column_letter(col)].width = 18

# ====== Format Headers (Row 1 & 2) ======
for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=5):
    for cell in row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = black_border

# ====== Merge State Column (A) ======
current_state = None
start_row = 3
max_row = ws.max_row
merged_state_ranges = []

for row in range(3, max_row + 2):
    cell = ws.cell(row=row, column=1)
    if cell.value and cell.value != current_state:
        if current_state is not None:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=row-1, end_column=1)
            merged_state_ranges.append((start_row, row-1))
        current_state = cell.value
        start_row = row
    elif row == max_row + 1:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=row-1, end_column=1)
        merged_state_ranges.append((start_row, row-1))

# ====== Format States Column (A) ======
for start_r, end_r in merged_state_ranges:
    for row in range(start_r, end_r + 1):
        cell = ws.cell(row=row, column=1)
        cell.fill = state_fill
        cell.font = state_font
        cell.alignment = center_align
        cell.border = black_border

# ====== Format Origins and Destinations (B to E) ======
for row in range(3, max_row + 1):
    for col in [2, 3, 4, 5]:
        cell = ws.cell(row=row, column=col)
        cell.fill = light_green_fill
        cell.font = white_bold_font
        cell.alignment = center_align
        cell.border = black_border

# ====== Freeze Panes below Header ======
ws.freeze_panes = "A3"

# ====== Save the File ======
wb.save(excel_path)
print(f"✅ Final version saved with white bold font + black borders in '{excel_path}'")
